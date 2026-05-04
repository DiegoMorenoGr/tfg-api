from io import BytesIO, StringIO
from datetime import date, datetime, time
from calendar import monthrange
import csv
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.core.database import Classification


def _month_range(year: int, month: int):
    start_date = datetime.combine(date(year, month, 1), time.min)
    end_date = datetime.combine(
        date(year, month, monthrange(year, month)[1]),
        time.max,
    )
    return start_date, end_date


def get_monthly_records(db: Session, year: int | None = None, month: int | None = None):
    today = date.today()

    if year is None:
        year = today.year

    if month is None:
        month = today.month

    start_date, end_date = _month_range(year, month)

    records = (
        db.query(Classification)
        .filter(Classification.timestamp >= start_date)
        .filter(Classification.timestamp <= end_date)
        .order_by(Classification.timestamp.asc())
        .all()
    )

    return records, year, month


def generate_monthly_csv(db: Session, year: int | None = None, month: int | None = None) -> BytesIO:
    records, year, month = get_monthly_records(db, year, month)

    text_stream = StringIO()
    writer = csv.writer(text_stream)

    writer.writerow([
        "subject",
        "sender",
        "category",
        "phishing_level",
    ])

    for r in records:
        writer.writerow([
            r.subject,
            r.sender,
            r.category,
            _get_phishing_level(r.phishing_score),
        ])

    output = BytesIO()
    output.write(text_stream.getvalue().encode("utf-8-sig"))
    output.seek(0)

    return output


def _get_phishing_level(score: float | None) -> str:
    if score is None:
        return "Desconocido"

    if score >= 0.75:
        return "Alto"
    elif score >= 0.40:
        return "Medio"
    else:
        return "Bajo"


def _week_label(timestamp: datetime) -> str:
    iso_year, iso_week, _ = timestamp.isocalendar()
    return f"{iso_year}-W{iso_week}"


def generate_monthly_excel(db: Session, year: int | None = None, month: int | None = None) -> BytesIO:
    records, year, month = get_monthly_records(db, year, month)

    wb = Workbook()

    # ==========================================================
    # HOJA 1: Reporte mensual
    # ==========================================================
    ws = wb.active
    ws.title = "Reporte mensual"

    ws["A1"] = f"Reporte mensual de correos - {month}/{year}"
    ws["A1"].font = Font(bold=True, size=14)

    # Tabla principal
    headers = [
        "Sujeto",
        "Remitente",
        "Categoría",
        "Nivel de phishing",
    ]

    start_row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True)

    category_counts = defaultdict(int)

    for row_index, r in enumerate(records, start=start_row + 1):
        category = r.category or "Sin categoría"
        category_counts[category] += 1

        ws.cell(row=row_index, column=1, value=r.subject)
        ws.cell(row=row_index, column=2, value=r.sender)
        ws.cell(row=row_index, column=3, value=category)
        ws.cell(row=row_index, column=4, value=_get_phishing_level(r.phishing_score))

    # Resumen por categoría debajo
    summary_start = start_row + len(records) + 3

    ws.cell(row=summary_start, column=1, value="Total por categoría")
    ws.cell(row=summary_start, column=1).font = Font(bold=True, size=12)

    ws.cell(row=summary_start + 1, column=1, value="Categoría")
    ws.cell(row=summary_start + 1, column=2, value="Total")
    ws.cell(row=summary_start + 1, column=1).font = Font(bold=True)
    ws.cell(row=summary_start + 1, column=2).font = Font(bold=True)

    for i, (category, total) in enumerate(
        sorted(category_counts.items(), key=lambda item: item[1], reverse=True),
        start=summary_start + 2,
    ):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=total)

    # Gráfica de barras por categoría
    if category_counts:
        chart = BarChart()
        chart.title = "Total de correos por categoría"
        chart.x_axis.title = "Categoría"
        chart.y_axis.title = "Número de correos"

        data = Reference(
            ws,
            min_col=2,
            min_row=summary_start + 1,
            max_row=summary_start + 1 + len(category_counts),
        )

        categories = Reference(
            ws,
            min_col=1,
            min_row=summary_start + 2,
            max_row=summary_start + 1 + len(category_counts),
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 14

        ws.add_chart(chart, f"D{summary_start}")

    # ==========================================================
    # HOJA 2: Evolución semanal
    # ==========================================================
    ws_weekly = wb.create_sheet("Correos por semana")

    ws_weekly["A1"] = "Evolución semanal de correos por categoría"
    ws_weekly["A1"].font = Font(bold=True, size=14)

    weekly_counts = defaultdict(lambda: defaultdict(int))
    categories = set()
    weeks = set()

    for r in records:
        if not r.timestamp:
            continue

        week = _week_label(r.timestamp)
        category = r.category or "Sin categoría"

        weekly_counts[week][category] += 1
        weeks.add(week)
        categories.add(category)

    sorted_weeks = sorted(weeks)
    sorted_categories = sorted(categories)

    table_start = 3

    ws_weekly.cell(row=table_start, column=1, value="Semana")
    ws_weekly.cell(row=table_start, column=1).font = Font(bold=True)

    for col_index, category in enumerate(sorted_categories, start=2):
        ws_weekly.cell(row=table_start, column=col_index, value=category)
        ws_weekly.cell(row=table_start, column=col_index).font = Font(bold=True)

    for row_index, week in enumerate(sorted_weeks, start=table_start + 1):
        ws_weekly.cell(row=row_index, column=1, value=week)

        for col_index, category in enumerate(sorted_categories, start=2):
            ws_weekly.cell(
                row=row_index,
                column=col_index,
                value=weekly_counts[week].get(category, 0),
            )

    # Gráfica de líneas
    if sorted_weeks and sorted_categories:
        line_chart = LineChart()
        line_chart.title = "Correos por semana y categoría"
        line_chart.x_axis.title = "Semana"
        line_chart.y_axis.title = "Número de correos"

        data = Reference(
            ws_weekly,
            min_col=2,
            max_col=1 + len(sorted_categories),
            min_row=table_start,
            max_row=table_start + len(sorted_weeks),
        )

        categories_ref = Reference(
            ws_weekly,
            min_col=1,
            min_row=table_start + 1,
            max_row=table_start + len(sorted_weeks),
        )

        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(categories_ref)
        line_chart.height = 10
        line_chart.width = 20

        ws_weekly.add_chart(line_chart, "A10")

    # Ajustar ancho de columnas
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output